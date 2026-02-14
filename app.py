# app.py
# Full app with:
# - 3 modes: Planning / Cumulative taper-aware / Match payslip (tax code + emergency)
# - Graphs
# - Excel export (data + Excel-native charts)
# - Inputs accept thousands separators (text inputs for £ fields)
# - Per-period table: whole £ only, commas, negatives in parentheses
# - Header hover tooltips (column_config help)
# - Column group toggles (sidebar)
# - FIX: Pension is based on PRE-salary-sacrifice pay (bonus only if elected)
# - NEW: Extra salary sacrifice can be applied in BONUS period only (March) and solver targets March-only

import re
from dataclasses import dataclass
from typing import Optional, List, Dict
from io import BytesIO

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


# ----------------------------
# Helpers
# ----------------------------
def clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(x, hi))


@dataclass
class TaxCodeInfo:
    raw: str
    country: str                 # "UK" | "S" | "C"
    is_emergency: bool           # W1/M1/X/NONCUM
    emergency_marker: Optional[str]
    code_type: str               # "standard" | "special" | "zero_allowance"
    special: Optional[str]       # BR/D0/D1/NT and variants
    allowance_annual: float      # can be negative for K codes
    suffix_letter: Optional[str] # L/M/N/T/K etc.


def decipher_tax_code(tax_code_raw: str) -> TaxCodeInfo:
    raw = (tax_code_raw or "").strip().upper()
    code = raw.replace(" ", "")

    is_emergency = False
    marker = None
    for m in ["W1", "M1", "X", "NONCUM"]:
        if code.endswith(m):
            is_emergency = True
            marker = m
            code = code[:-len(m)]
            break

    country = "UK"
    if code.startswith("S"):
        country = "S"
        code = code[1:]
    elif code.startswith("C"):
        country = "C"
        code = code[1:]

    special_codes = {
        "BR", "D0", "D1", "NT",
        "SBR", "SD0", "SD1", "SD2", "SD3",
        "CBR", "CD0", "CD1"
    }
    if code in special_codes:
        return TaxCodeInfo(raw, country, is_emergency, marker, "special", code, 0.0, None)

    if code == "0T":
        return TaxCodeInfo(raw, country, is_emergency, marker, "zero_allowance", None, 0.0, "T")

    m_k = re.fullmatch(r"K(\d+)", code)
    if m_k:
        n = int(m_k.group(1))
        return TaxCodeInfo(raw, country, is_emergency, marker, "standard", None, -10.0 * n, "K")

    m_std = re.fullmatch(r"(\d+)([A-Z])", code)
    if m_std:
        n = int(m_std.group(1))
        letter = m_std.group(2)
        return TaxCodeInfo(raw, country, is_emergency, marker, "standard", None, 10.0 * n, letter)

    return TaxCodeInfo(raw, country, is_emergency, marker, "standard", None, 12570.0, "L")


def tapered_personal_allowance(annualised_ani: float, base_pa: float, taper_threshold: float) -> float:
    if annualised_ani <= taper_threshold:
        return base_pa
    reduction = (annualised_ani - taper_threshold) / 2.0
    return max(0.0, base_pa - reduction)


def build_period_labels(periods: int) -> List[str]:
    if periods == 12:
        return ["April","May","June","July","August","September","October","November","December","January","February","March"]
    return [f"P{i}" for i in range(1, periods + 1)]


def compute_tax_cumulative(
    cum_adj: float,
    idx: int,
    allowance_annual: float,
    basic_limit: float,
    higher_limit: float,
    r_basic: float,
    r_higher: float,
    r_add: float,
    periods: int
) -> Dict[str, float]:
    allowance_to_date = allowance_annual * idx / periods
    taxable_to_date = max(0.0, cum_adj - allowance_to_date)

    basic_band_to_date = max(0.0, (basic_limit - max(0.0, allowance_annual)) * idx / periods)
    higher_band_to_date = max(0.0, (higher_limit - basic_limit) * idx / periods)

    cum_b = clamp(taxable_to_date, 0.0, basic_band_to_date) * r_basic
    cum_h = clamp(taxable_to_date - basic_band_to_date, 0.0, higher_band_to_date) * r_higher
    cum_a = max(0.0, taxable_to_date - (basic_band_to_date + higher_band_to_date)) * r_add
    return {"b": cum_b, "h": cum_h, "a": cum_a, "t": cum_b + cum_h + cum_a}


# ----------------------------
# Input formatting: thousands-separated money inputs (text)
# ----------------------------
def money_input(label: str, default: int = 0, key: Optional[str] = None) -> int:
    raw = st.text_input(label, value=f"{int(default):,}", key=key)
    try:
        cleaned = raw.replace(",", "").replace("£", "").replace(" ", "")
        if cleaned == "":
            return 0
        return int(round(float(cleaned), 0))
    except Exception:
        return 0


# ----------------------------
# Display formatting: whole £, commas, negatives in parentheses
# ----------------------------
def fmt_money_str(x) -> str:
    if pd.isna(x):
        return ""
    v = int(round(float(x), 0))
    return f"({abs(v):,})" if v < 0 else f"{v:,}"


def fmt_int_str(x) -> str:
    if pd.isna(x):
        return ""
    return f"{int(round(float(x), 0)):,}"


# ----------------------------
# Column groups for toggles
# ----------------------------
COLUMN_GROUPS = {
    "Core pay": [
        "Period", "Gross Salary", "Gross Bonus", "Total Gross",
        "Salary Sacrifice (Total)", "Pension", "Adjusted Taxable"
    ],
    "Income tax (summary)": [
        "Total Income Tax"
    ],
    "Income tax (by band)": [
        "Income Tax - Basic", "Income Tax - Higher", "Income Tax - Additional"
    ],
    "National Insurance": [
        "NI - Main", "NI - Upper", "Total NI"
    ],
    "Net & cumulative": [
        "Net Take Home", "Cum Net", "Cum Tax", "Cum NI"
    ],
    "Allowances / taper": [
        "Annualised ANI (est.)",
        "Effective PA (annual est.)",
        "Effective PA (annual)",
        "Tax Code Allowance (annual)",
    ],
    "Tax code": [
        "Tax Code", "Emergency / Non-cum"
    ],
}


# ----------------------------
# Excel export (data + charts)
# ----------------------------
def build_excel_export(df_export: pd.DataFrame, summary: Dict[str, float]) -> BytesIO:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "Metric"
    ws_sum["B1"] = "Value"
    r = 2
    for k, v in summary.items():
        ws_sum[f"A{r}"] = k
        ws_sum[f"B{r}"] = float(v)
        r += 1

    ws = wb.create_sheet("Data")
    for row in dataframe_to_rows(df_export, index=False, header=True):
        ws.append(row)

    ws.freeze_panes = "A2"
    ws_sum.freeze_panes = "A2"

    n_rows = ws.max_row
    n_cols = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, n_cols + 1)]
    col = {h: i+1 for i, h in enumerate(headers) if h is not None}

    cat_ref = Reference(ws, min_col=col["Period"], min_row=2, max_row=n_rows) if "Period" in col else None

    # Line chart: Net vs Tax vs NI
    line = LineChart()
    line.title = "Net vs Tax vs NI"
    line.y_axis.title = "£"
    line.x_axis.title = "Period"

    series_cols = [c for c in ["Net Take Home", "Total Income Tax", "Total NI"] if c in col]
    if cat_ref and series_cols:
        data_ref = Reference(ws, min_col=col[series_cols[0]], min_row=1,
                             max_col=col[series_cols[-1]], max_row=n_rows)
        line.add_data(data_ref, titles_from_data=True)
        line.set_categories(cat_ref)
        line.height = 10
        line.width = 24
        ws.add_chart(line, "R2")

    # Stacked bar chart: tax band split
    bars = BarChart()
    bars.type = "col"
    bars.grouping = "stacked"
    bars.title = "Income Tax split by bracket"
    bars.y_axis.title = "£"
    bars.x_axis.title = "Period"
    bars.dataLabels = DataLabelList()
    bars.dataLabels.showVal = False

    tax_cols = [c for c in ["Income Tax - Basic", "Income Tax - Higher", "Income Tax - Additional"] if c in col]
    if cat_ref and tax_cols:
        tax_ref = Reference(ws, min_col=col[tax_cols[0]], min_row=1,
                            max_col=col[tax_cols[-1]], max_row=n_rows)
        bars.add_data(tax_ref, titles_from_data=True)
        bars.set_categories(cat_ref)
        bars.height = 10
        bars.width = 24
        ws.add_chart(bars, "R22")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="UK PAYE Tracker + Taper", layout="wide")
st.title("UK Monthly Income Tracker (Planning / Cumulative / Match Payslip)")

with st.sidebar:
    st.header("Mode")
    mode = st.selectbox(
        "Calculation mode",
        ["Planning (monthlyised)", "Cumulative PAYE (approx, taper-aware)", "Match payslip (tax code + frequency)"],
        index=2
    )

    st.divider()
    st.header("Pay frequency")
    freq = st.selectbox("Pay frequency", ["Monthly", "Weekly", "Fortnightly", "4-weekly"], index=0)
    periods_map = {"Monthly": 12, "Weekly": 52, "Fortnightly": 26, "4-weekly": 13}
    periods = periods_map[freq]
    labels = build_period_labels(periods)

    st.divider()
    st.header("Income inputs")
    salary = money_input("Annual salary (£)", 35000, key="salary")
    bonus = money_input("Bonus amount (£)", 0, key="bonus")

    if periods == 12:
        bonus_label = st.selectbox("Bonus month", labels, index=11)
        bonus_period = labels.index(bonus_label) + 1
    else:
        bonus_period = st.number_input(
            f"Bonus period number (1..{periods})",
            value=min(periods, 12),
            step=1,
            min_value=1,
            max_value=periods
        )

    st.divider()
    st.header("Match payslip inputs")
    tax_code_raw = st.text_input(
        "Tax code (e.g., 1257L, 1257L M1, 0T, BR, D0, D1, NT, K500, S1257L)",
        value="1257L"
    )

    st.divider()
    st.header("Salary sacrifice & pension")
    fixed_ss_monthly = money_input("Fixed salary sacrifice per month (£)", 0, key="fixed_ss")
    extra_ss_annual = money_input("Extra salary sacrifice in bonus period (£)", 0, key="extra_ss")

    extra_ss_in_bonus_only = st.toggle(
        "Apply extra salary sacrifice in bonus period only",
        value=True
    )

    pension_rate = st.number_input(
        "Pension rate (%)",
        value=0.0,
        step=0.5,
        min_value=0.0,
        max_value=100.0
    ) / 100.0

    apply_fixed_ss_in_bonus = st.toggle("Apply fixed SS in bonus period too", value=True)

    # default OFF = bonus doesn't affect pension unless you elect it (taper lever)
    apply_pension_to_bonus = st.toggle("Apply pension on bonus too", value=False)

    st.divider()
    st.header("Per-period table columns")
    visible_groups: Dict[str, bool] = {}
    for group, _cols in COLUMN_GROUPS.items():
        default_on = group in ["Core pay", "Income tax (summary)", "National Insurance", "Net & cumulative"]
        visible_groups[group] = st.checkbox(group, value=default_on)

    st.divider()
    st.header("Income Tax parameters")
    PA_base = money_input("Personal Allowance (base) (£)", 12570, key="pa_base")
    taper_threshold = money_input("£100k taper threshold (£)", 100000, key="taper_threshold")
    basic_limit = money_input("Basic rate limit (£)", 50270, key="basic_limit")
    higher_limit = money_input("Higher rate limit (£)", 125140, key="higher_limit")

    r_basic = st.number_input("Basic rate", value=0.20, step=0.01, min_value=0.0, max_value=1.0)
    r_higher = st.number_input("Higher rate", value=0.40, step=0.01, min_value=0.0, max_value=1.0)
    r_add = st.number_input("Additional rate", value=0.45, step=0.01, min_value=0.0, max_value=1.0)

    st.divider()
    st.header("National Insurance")
    NI_PT = money_input("NI Primary Threshold (annual £)", 12570, key="ni_pt")
    NI_UEL = money_input("NI Upper Earnings Limit (annual £)", 50270, key="ni_uel")
    NI_main = st.number_input("NI main rate", value=0.08, step=0.01, min_value=0.0, max_value=1.0)
    NI_upper = st.number_input("NI upper rate", value=0.02, step=0.01, min_value=0.0, max_value=1.0)

    st.divider()
    st.header("Visuals")
    show_cumulative_lines = st.toggle("Show cumulative lines", value=True)


# ----------------------------
# Build per-period table
# - Pension is based on PRE-salary-sacrifice pay
# - Extra salary sacrifice can be bonus-period only
# ----------------------------
salary_per_period = salary / periods
fixed_ss_per_period = fixed_ss_monthly * 12.0 / periods  # monthly -> per-period
extra_ss_spread = extra_ss_annual / periods  # used if not bonus-only

rows = []
for i, lab in enumerate(labels, start=1):
    gross_salary = salary_per_period
    gross_bonus = bonus if i == bonus_period else 0.0
    total_gross = gross_salary + gross_bonus

    fixed_ss = fixed_ss_per_period if (apply_fixed_ss_in_bonus or i != bonus_period) else 0.0

    if extra_ss_in_bonus_only:
        extra_ss_this_period = float(extra_ss_annual) if i == bonus_period else 0.0
    else:
        extra_ss_this_period = float(extra_ss_spread)

    total_ss = fixed_ss + extra_ss_this_period

    pension_base = gross_salary
    if apply_pension_to_bonus:
        pension_base += gross_bonus
    pension = pension_base * pension_rate

    adj_taxable = total_gross - total_ss - pension

    rows.append([lab, gross_salary, gross_bonus, total_gross, total_ss, pension, adj_taxable])

df = pd.DataFrame(rows, columns=[
    "Period", "Gross Salary", "Gross Bonus", "Total Gross",
    "Salary Sacrifice (Total)", "Pension", "Adjusted Taxable"
])

# ----------------------------
# Income tax
# ----------------------------
df["Income Tax - Basic"] = 0.0
df["Income Tax - Higher"] = 0.0
df["Income Tax - Additional"] = 0.0
df["Total Income Tax"] = 0.0

if mode == "Planning (monthlyised)":
    annual_adj = float(df["Adjusted Taxable"].sum())
    PA_eff = tapered_personal_allowance(annual_adj, float(PA_base), float(taper_threshold))

    basic_band_per = max(0.0, (float(basic_limit) - PA_eff) / periods)
    higher_band_per = max(0.0, (float(higher_limit) - float(basic_limit)) / periods)
    add_threshold_per = float(higher_limit) / periods

    b_list, h_list, a_list, t_list = [], [], [], []
    for x in df["Adjusted Taxable"]:
        x = float(x)
        b = clamp(x, 0.0, basic_band_per) * r_basic
        h = clamp(x - basic_band_per, 0.0, higher_band_per) * r_higher
        a = max(0.0, x - add_threshold_per) * r_add
        b_list.append(b); h_list.append(h); a_list.append(a); t_list.append(b + h + a)

    df["Effective PA (annual)"] = PA_eff
    df["Income Tax - Basic"] = b_list
    df["Income Tax - Higher"] = h_list
    df["Income Tax - Additional"] = a_list
    df["Total Income Tax"] = t_list

elif mode == "Cumulative PAYE (approx, taper-aware)":
    cum_adj = 0.0
    cum_tax_prev = 0.0
    cum_b_prev = cum_h_prev = cum_a_prev = 0.0

    b_d, h_d, a_d, t_d, pa_eff_list, ani_ann_list = [], [], [], [], [], []

    for idx, x in enumerate(df["Adjusted Taxable"], start=1):
        cum_adj += float(x)
        annualised_ani = (cum_adj * periods) / idx
        ani_ann_list.append(annualised_ani)

        PA_eff = tapered_personal_allowance(annualised_ani, float(PA_base), float(taper_threshold))
        pa_eff_list.append(PA_eff)

        cum_components = compute_tax_cumulative(
            cum_adj=cum_adj,
            idx=idx,
            allowance_annual=PA_eff,
            basic_limit=float(basic_limit),
            higher_limit=float(higher_limit),
            r_basic=r_basic,
            r_higher=r_higher,
            r_add=r_add,
            periods=periods
        )

        t_d.append(cum_components["t"] - cum_tax_prev)
        b_d.append(cum_components["b"] - cum_b_prev)
        h_d.append(cum_components["h"] - cum_h_prev)
        a_d.append(cum_components["a"] - cum_a_prev)

        cum_tax_prev = cum_components["t"]
        cum_b_prev = cum_components["b"]
        cum_h_prev = cum_components["h"]
        cum_a_prev = cum_components["a"]

    df["Annualised ANI (est.)"] = ani_ann_list
    df["Effective PA (annual est.)"] = pa_eff_list
    df["Income Tax - Basic"] = b_d
    df["Income Tax - Higher"] = h_d
    df["Income Tax - Additional"] = a_d
    df["Total Income Tax"] = t_d

else:
    tc = decipher_tax_code(tax_code_raw)
    st.sidebar.caption(
        f"Tax code interpretation: {tc.country} | {tc.code_type} | {tc.special or ''} | "
        f"Allowance £{tc.allowance_annual:,.0f} | "
        f"{'Emergency ' + (tc.emergency_marker or '') if tc.is_emergency else 'Cumulative'}"
    )

    cum_adj = 0.0
    cum_tax_prev = 0.0
    cum_b_prev = cum_h_prev = cum_a_prev = 0.0

    b_d, h_d, a_d, t_d = [], [], [], []
    for idx, x in enumerate(df["Adjusted Taxable"], start=1):
        x = float(x)

        if tc.is_emergency:
            if tc.code_type == "special":
                if tc.special.endswith("NT"):
                    b = h = a = t = 0.0
                elif tc.special.endswith("BR"):
                    b = x * r_basic; h = a = 0.0; t = b
                elif tc.special.endswith("D0"):
                    h = x * r_higher; b = a = 0.0; t = h
                elif tc.special.endswith("D1"):
                    a = x * r_add; b = h = 0.0; t = a
                else:
                    b = h = a = t = 0.0
            else:
                allowance_per = tc.allowance_annual / periods
                taxable = max(0.0, x - allowance_per)

                basic_band_per = max(0.0, (float(basic_limit) - max(0.0, tc.allowance_annual)) / periods)
                higher_band_per = max(0.0, (float(higher_limit) - float(basic_limit)) / periods)

                b = clamp(taxable, 0.0, basic_band_per) * r_basic
                h = clamp(taxable - basic_band_per, 0.0, higher_band_per) * r_higher
                a = max(0.0, taxable - (basic_band_per + higher_band_per)) * r_add
                t = b + h + a

            b_d.append(b); h_d.append(h); a_d.append(a); t_d.append(t)

        else:
            cum_adj += x

            if tc.code_type == "special":
                if tc.special.endswith("NT"):
                    cum_b = cum_h = cum_a = cum_t = 0.0
                elif tc.special.endswith("BR"):
                    cum_t = cum_adj * r_basic
                    cum_b, cum_h, cum_a = cum_t, 0.0, 0.0
                elif tc.special.endswith("D0"):
                    cum_t = cum_adj * r_higher
                    cum_b, cum_h, cum_a = 0.0, cum_t, 0.0
                elif tc.special.endswith("D1"):
                    cum_t = cum_adj * r_add
                    cum_b, cum_h, cum_a = 0.0, 0.0, cum_t
                else:
                    cum_b = cum_h = cum_a = cum_t = 0.0
            else:
                comp = compute_tax_cumulative(
                    cum_adj=cum_adj,
                    idx=idx,
                    allowance_annual=tc.allowance_annual,
                    basic_limit=float(basic_limit),
                    higher_limit=float(higher_limit),
                    r_basic=r_basic,
                    r_higher=r_higher,
                    r_add=r_add,
                    periods=periods
                )
                cum_b, cum_h, cum_a, cum_t = comp["b"], comp["h"], comp["a"], comp["t"]

            t_d.append(cum_t - cum_tax_prev)
            b_d.append(cum_b - cum_b_prev)
            h_d.append(cum_h - cum_h_prev)
            a_d.append(cum_a - cum_a_prev)

            cum_tax_prev = cum_t
            cum_b_prev = cum_b
            cum_h_prev = cum_h
            cum_a_prev = cum_a

    df["Income Tax - Basic"] = b_d
    df["Income Tax - Higher"] = h_d
    df["Income Tax - Additional"] = a_d
    df["Total Income Tax"] = t_d
    df["Tax Code"] = tc.raw
    df["Tax Code Allowance (annual)"] = tc.allowance_annual
    df["Emergency / Non-cum"] = tc.is_emergency


# ----------------------------
# National Insurance
# ----------------------------
PT_p = float(NI_PT) / periods
UEL_p = float(NI_UEL) / periods

ni_main_list, ni_upper_list = [], []
for x in df["Adjusted Taxable"]:
    x = float(x)
    ni_main_amt = max(0.0, min(x, UEL_p) - PT_p) * NI_main
    ni_upper_amt = max(0.0, x - UEL_p) * NI_upper
    ni_main_list.append(ni_main_amt)
    ni_upper_list.append(ni_upper_amt)

df["NI - Main"] = ni_main_list
df["NI - Upper"] = ni_upper_list
df["Total NI"] = df[["NI - Main", "NI - Upper"]].sum(axis=1)

df["Net Take Home"] = df["Adjusted Taxable"] - df["Total Income Tax"] - df["Total NI"]
df["Cum Net"] = df["Net Take Home"].cumsum()
df["Cum Tax"] = df["Total Income Tax"].cumsum()
df["Cum NI"] = df["Total NI"].cumsum()

# ----------------------------
# Summary + Export
# ----------------------------
annual_gross = float(df["Total Gross"].sum())
annual_sacrifice = float(df["Salary Sacrifice (Total)"].sum())
annual_pension = float(df["Pension"].sum())
annual_ani = float(df["Adjusted Taxable"].sum())
annual_tax = float(df["Total Income Tax"].sum())
annual_ni = float(df["Total NI"].sum())
annual_net = float(df["Net Take Home"].sum())

summary = {
    "Total gross (£)": annual_gross,
    "Total sacrifice (£)": annual_sacrifice,
    "Total pension (£)": annual_pension,
    "Adjusted net income (£)": annual_ani,
    "Total income tax (£)": annual_tax,
    "Total NI (£)": annual_ni,
    "Total net take-home (£)": annual_net,
}

excel_bytes = build_excel_export(df, summary)

# ----------------------------
# UI
# ----------------------------
tab1, tab2, tab3, tab4 = st.tabs(["Main", "Graphs", "Export", "Taper Solver"])

money_cols = [
    "Gross Salary","Gross Bonus","Total Gross",
    "Salary Sacrifice (Total)","Pension","Adjusted Taxable",
    "Income Tax - Basic","Income Tax - Higher","Income Tax - Additional","Total Income Tax",
    "NI - Main","NI - Upper","Total NI",
    "Net Take Home","Cum Net","Cum Tax","Cum NI",
]
int_like_cols = [
    "Annualised ANI (est.)", "Effective PA (annual est.)", "Effective PA (annual)",
    "Tax Code Allowance (annual)",
]

col_help = {
    "Period": "Pay period label (e.g., April…March).",
    "Gross Salary": "Base salary paid in that period (before deductions).",
    "Gross Bonus": "Bonus paid in that period (0 in non-bonus periods).",
    "Total Gross": "Gross Salary + Gross Bonus.",
    "Salary Sacrifice (Total)": "Pre-tax salary sacrifice deducted this period.",
    "Pension": "Employee pension contribution deducted this period (pre-sacrifice base).",
    "Adjusted Taxable": "Total Gross - Salary Sacrifice - Pension (used for PAYE/NI in this model).",
    "Income Tax - Basic": "Income tax charged at basic rate in this period.",
    "Income Tax - Higher": "Income tax charged at higher rate in this period.",
    "Income Tax - Additional": "Income tax charged at additional rate in this period.",
    "Total Income Tax": "Total income tax in this period (sum of brackets).",
    "NI - Main": "Employee NI at main rate in this period.",
    "NI - Upper": "Employee NI at upper rate in this period.",
    "Total NI": "Total employee NI in this period.",
    "Net Take Home": "Adjusted Taxable - Total Income Tax - Total NI.",
    "Cum Net": "Cumulative net take-home to date.",
    "Cum Tax": "Cumulative income tax to date.",
    "Cum NI": "Cumulative NI to date.",
    "Tax Code": "Tax code used in Match Payslip mode.",
    "Tax Code Allowance (annual)": "Annual allowance implied by the tax code (can be negative for K codes).",
    "Emergency / Non-cum": "If true, payslip uses W1/M1/X/NONCUM non-cumulative taxation.",
    "Annualised ANI (est.)": "Estimated annualised adjusted net income based on year-to-date.",
    "Effective PA (annual est.)": "Estimated personal allowance after £100k taper (cumulative mode).",
    "Effective PA (annual)": "Personal allowance used in Planning mode (after taper).",
}

with tab1:
    st.subheader("Per-period breakdown")

    df_display = df.copy()

    for c in money_cols:
        if c in df_display.columns:
            df_display[c] = df_display[c].map(fmt_money_str)
    for c in int_like_cols:
        if c in df_display.columns:
            df_display[c] = df_display[c].map(fmt_int_str)

    visible_columns: List[str] = []
    for group, enabled in visible_groups.items():
        if enabled:
            for c in COLUMN_GROUPS[group]:
                if c in df_display.columns and c not in visible_columns:
                    visible_columns.append(c)

    if not visible_columns:
        visible_columns = ["Period"] if "Period" in df_display.columns else list(df_display.columns[:1])

    df_visible = df_display[visible_columns]
    column_config = {c: st.column_config.TextColumn(help=col_help.get(c, "")) for c in df_visible.columns}

    st.dataframe(df_visible, use_container_width=True, column_config=column_config)

    st.subheader("Summary")
    summary_items = [
        ("Total gross (£)", annual_gross),
        ("Total sacrifice (£)", annual_sacrifice),
        ("Total pension (£)", annual_pension),
        ("Adjusted net income (£)", annual_ani),
        ("Total income tax (£)", annual_tax),
        ("Total NI (£)", annual_ni),
        ("Total net take-home (£)", annual_net),
    ]
    summary_df = pd.DataFrame(summary_items, columns=["Metric", "Value (£)"])
    summary_df["Value (£)"] = summary_df["Value (£)"].map(fmt_money_str)
    st.table(summary_df)

    st.caption("Location context: xxxxxxxxx")

with tab2:
    st.subheader("Visualisations")

    fig1 = plt.figure()
    plt.plot(df["Period"], df["Net Take Home"], marker="o")
    plt.plot(df["Period"], df["Total Income Tax"], marker="o")
    plt.plot(df["Period"], df["Total NI"], marker="o")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("£")
    plt.title("Per-period: Net vs Tax vs NI")
    plt.tight_layout()
    st.pyplot(fig1)

    if show_cumulative_lines:
        fig2 = plt.figure()
        plt.plot(df["Period"], df["Cum Net"], marker="o")
        plt.plot(df["Period"], df["Cum Tax"], marker="o")
        plt.plot(df["Period"], df["Cum NI"], marker="o")
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("£ (cumulative)")
        plt.title("Cumulative: Net vs Tax vs NI")
        plt.tight_layout()
        st.pyplot(fig2)

    fig3 = plt.figure()
    b = df["Income Tax - Basic"]
    h = df["Income Tax - Higher"]
    a = df["Income Tax - Additional"]
    plt.bar(df["Period"], b, label="Basic")
    plt.bar(df["Period"], h, bottom=b, label="Higher")
    plt.bar(df["Period"], a, bottom=b + h, label="Additional")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("£")
    plt.title("Income Tax split by bracket (stacked)")
    plt.legend()
    plt.tight_layout()
    st.pyplot(fig3)

with tab3:
    st.subheader("Export")
    st.download_button(
        label="Download Excel (.xlsx)",
        data=excel_bytes,
        file_name="uk_income_tracker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with tab4:
    st.subheader("Taper Solver (£100k adjusted net income)")

    annual_gross_solver = float(salary + bonus)

    fixed_ss_annual = float(fixed_ss_monthly * 12.0)
    current_total_ss_without_extra = float(fixed_ss_annual)

    pension_base_annual = float(salary)
    if apply_pension_to_bonus:
        pension_base_annual += float(bonus)
    pension_annual = pension_base_annual * float(pension_rate)

    current_ani_no_extra = annual_gross_solver - current_total_ss_without_extra - pension_annual

    required_march_extra_ss = max(0.0, current_ani_no_extra - float(taper_threshold))

    st.write(f"Current estimated adjusted net income (before extra bonus-period SS): **£{current_ani_no_extra:,.0f}**")
    st.write(f"Taper threshold: **£{float(taper_threshold):,.0f}**")

    st.metric("Extra salary sacrifice needed in bonus period (£)", f"{required_march_extra_ss:,.0f}")

    if extra_ss_in_bonus_only:
        st.caption("Extra salary sacrifice input is treated as a bonus-period-only amount.")
    else:
        st.caption("Extra salary sacrifice input is currently spread across periods (toggle it to bonus-period-only).")


